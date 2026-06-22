from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Jugadores"

headers = [
    "Nombre completo", "Género", "Año universitario", "Altura", "Peso",
    "Posición", "Equipo/Club en Argentina (pasado)", "Liga en Argentina",
    "Universidad / College (USA)", "División (NCAA D1/D2/NAIA)",
    "Temporada actual - Partidos jugados", "Temporada actual - Minutos prom.",
    "Temporada actual - Puntos prom.", "Temporada actual - Rebotes prom.",
    "Temporada actual - Asistencias prom.", "Estado", "Fuente / Link", "Notas"
]

data = [
["Juan Sebastián Gorosito","M","N/D","6-1","190 lbs","Base / Escolta","N/D","N/D","San Diego","D1","","","","","","Activo","","Verificado en lista original"],
["Santiago Trouet","M","Redshirt Sophomore","6-10","218 lbs","Ala-pívot","Estudiantes de Madrid (España)","Liga Endesa (España)","Arizona State","D1","","","","","","Activo","",""],
["Bautista Giralt","M","Redshirt Sophomore","7-0","235 lbs","Pívot","Obras Basket","Liga Argentina","Albany","D1","","","","","","Activo","",""],
["Tyler Kropp","M","Sophomore","6-9","230 lbs","Ala-pívot","Olentangy Liberty HS (USA)","N/A","Washington State","D1","","","","","","Activo","",""],
["Thiago Sucatzky","M","N/D","5-10","190 lbs","Base","N/D","N/D","FIU","D1","","","","","","Activo","",""],
["Marcus Adams Jr.","M","Redshirt Sophomore","6-8","N/D","Alero","Narbonne HS (USA)","N/A","Arizona State","D1","","","","","","Activo","",""],
["Felipe Minzer","M","Freshman","6-6","N/D","Escolta / Alero","Zaragoza (España)","Liga Endesa (España)","Austin Peay","D1","","","","","","Activo","",""],
["Lucas Mercandino","M","N/D","6-6","200 lbs","Alero","N/D","N/D","Lubbock Christian","NAIA","","","","","","Activo","",""],
["Ignacio García","M","N/D","6-4","200 lbs","Base","N/D","N/D","Lubbock Christian","NAIA","","","","","","Activo","",""],
["Facundo Aranda","M","N/D","6-0","175 lbs","Base","N/D","N/D","Biola","D2","","","","","","Activo","",""],
["Bautista Rodríguez","M","N/D","6-7","210 lbs","Alero","N/D","N/D","Indiana Univ. of Pennsylvania","D2","","","","","","Activo","",""],
["Guido Wanschelbaum","M","N/D","6-9","210 lbs","Pívot","N/D","N/D","South Dakota Mines","D2","","","","","","Activo","",""],
["Matías Zanotto","M","N/D","6-9","205 lbs","Ala-pívot","N/D","N/D","Augusta","D2","","","","","","Activo","",""],
["Tobías Lábaque","M","N/D","6-3","188 lbs","Escolta","N/D","N/D","Lynchburg","D3","","","","","","Activo","",""],
["Bernardo Zappia","M","N/D","6-7","N/D","Pívot","N/D","N/D","Whittier","D3","","","","","","Activo","",""],
["Máximo Adams","M","High School (futuro college)","N/D","N/D","Alero","Sierra Canyon HS (USA)","N/A","North Carolina (compromiso)","D1","","","","","","Comprometido / aún no debuta","",""],
["Paula López","F","Senior","5-8","N/D","Base / Escolta","San Blas Alicante (España)","Liga Femenina (España)","Detroit Mercy","D1","","","","","","Activa","",""],
["Sol Depetris","F","Freshman","5-7","N/D","Base","Unión Florida","Liga Argentina Fem.","Bowling Green","D1","","","","","","Activa","",""],
["Brenda Fontana","F","Senior","6-1","N/D","Ala-pívot","Vélez Sarsfield","Liga Argentina Fem.","Old Dominion","D1","","","","","","Activa","",""],
["Natalia Tondi","F","Graduada","5-7","N/D","Base","Unión Florida","Liga Argentina Fem.","Virginia Commonwealth","D1","","","","","","Activa","",""],
["Tiziana Huici","F","N/D","5-8","N/D","Base / Escolta","Tomás de Rocamora","Liga Argentina Fem.","Queens (NC)","D2","","","","","","Activa","",""],
["Georgina Buzzetti","F","N/D","N/D","N/D","N/D","Lanús","Liga Argentina Fem.","Florida South Western","NJCAA","","","","","","Activa","",""],
["Delfina Cergneaux","F","N/D","N/D","N/D","N/D","N/D","N/D","Florida South Western","NJCAA","","","","","","Activa","",""],
["Julia Paoletta","F","N/D","N/D","N/D","N/D","N/D","N/D","Newberry","D2","","","","","","Activa","",""],
["Sol Castro","F","N/D","N/D","N/D","N/D","N/D","N/D","San Francisco","N/D","","","","","","Activa","",""],
["Francisco Farabello","M","N/D (ex TCU)","N/D","N/D","Base","N/D","N/D","Creighton","D1","","","","","","Activo","https://www.basquetplus.com/ncaa-estados-unidos-arranca-temporada-argentinos-francisco-farabello-francisco-caffaro-rafael-martinez-lucas-mercandino-juanse-gorosito-felipe-palazzo","Hijo de Daniel Farabello, jugó antes en TCU"],
["Felipe Palazzo","M","N/D","N/D","N/D","Pívot / Ala-pívot","Tucumán Básquet","Liga Argentina","Oregon State","D1","","","","","","Activo","https://www.basquetplus.com/ncaa-estados-unidos-arranca-temporada-argentinos-francisco-farabello-francisco-caffaro-rafael-martinez-lucas-mercandino-juanse-gorosito-felipe-palazzo","Saltó directo de Liga Argentina a NCAA"],
["Stéfano Alesso","M","N/D","7-1 (2.15m)","N/D","Pívot","N/D (formado en Salta / Oberá)","N/D","Stetson","D1","","","","","","Activo","https://basquetplus.com/ncaa-stefano-alesso-universidad-stetson-historia-pivote-gigante-salta-italia-obera-promesa-seleccion-argentina","Nacido en La Paz (Bolivia), criado en Salta"],
["Alonso Easterling","M","N/D","6-11 (2.08m)","N/D","Ala-pívot / Pívot","N/D","N/A (vive en USA desde los 4 años)","Canisius","D1","","","","","","Activo / a confirmar universidad exacta","https://basquetplus.com/ncaa-estados-unidos-alonso-easterling-nuevo-jugador-universidad-gardner-webb","Hijo de Stanley Easterling; fuentes mencionan también Gardner-Webb, verificar"],
["Máximo Milovich","M","N/D (transferido)","N/D","N/D","N/D","N/D","N/D","UC San Diego","D1","","","","","","Activo (transfer)","https://www.basquetplus.com/ncaa-maximo-milovich-division-I-uc-san-diego-transferencia","Llegó por transferencia a Division I"],
["Francisco Cáffaro","M","N/D (carrera NCAA finalizada o por confirmar)","N/D","N/D","Pívot","N/D","N/D","Virginia / Santa Clara","D1","","","","","","A confirmar si sigue activo en NCAA","https://en.wikipedia.org/wiki/Francisco_C%C3%A1ffaro","Campeón NCAA 2019 con Virginia; revisar si ya pasó a profesional"],
["Juan Fernández","M","N/D (recién comprometido)","N/D","N/D","Base / Escolta","N/D","N/D","South Carolina (desde 2026-27)","D1","","","","","","Comprometido, aún no debuta en NCAA","https://basquetplus.com/ncaa-nil-juan-fernandez-girona-south-carolina-dinero-lee-aaliya","Venía de ser profesional en Girona (España) — caso atípico, no amateur tradicional"],
]

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="1F4E78")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for r, row in enumerate(data, 2):
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val)

widths = [22,8,22,10,10,18,28,20,26,12,14,14,12,14,16,28,40,40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

for font_name in ["Jugadores"]:
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", bold=cell.font.bold, color=cell.font.color)

wb.save("Jugadores_Argentinos_USA.xlsx")
print("saved")
