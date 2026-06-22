from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Jugadores"

headers = [
    "Nombre completo", "Género", "Nivel/Tier de certeza", "Año/Clase",
    "Posición", "Origen en Argentina", "Equipo/Club en Argentina (pasado)",
    "Institución actual en USA", "División", "Estado al 22/06/2026",
    "Fecha de confirmación", "Fuente", "Notas"
]

TIER1 = "Confirmado 2026-27"
TIER2 = "Prospecto / Prep -> NCAA"
TIER3 = "Última confirmación oficial 2025-26"
TIER4 = "Pendiente de verificar (no en informe estricto)"
TIER5 = "Descartado / cautela (no cumple estándar estricto)"

data = [
["Santiago Trouet","M",TIER1,"N/D","Ala-pívot","Buenos Aires","Estudiantes de Madrid (España) / Arizona State (2025-26)","Ole Miss","NCAA D-I","Confirmado para 2026-27 tras transfer","01/05/2026","https://thesundevils.com/sports/mens-basketball/roster/player/santiago-trouet ; https://247sports.com/player/santiago-trouet-46157048/college-334115/","Transferencia confirmada por 247Sports el 1/5/2026"],
["Tyler Kropp","M",TIER1,"Sophomore","Ala-pívot","N/D","Olentangy Liberty HS (USA)","Washington State","NCAA D-I","Confirmado en roster 2026-27","N/D","https://wsucougars.com/sports/mens-basketball/roster/tyler-kropp/15512","Vinculado a la selección argentina por FIBA"],
["Thiago Sucatzky","M",TIER1,"N/D","Base","N/D","FIU (2025-26)","Stetson","NCAA D-I","Confirmado por transferencia de primavera 2026","11/04/2026","https://247sports.com/player/thiago-sucatzky-46166121/","Stetson MBB publicó bienvenida oficial"],
["Georgi Buzzetti","F",TIER1,"N/D","N/D","Buenos Aires","Lanús / Florida SouthWestern (2025-26)","North Florida","NCAA D-I","Confirmada en roster 2026-27","N/D","https://unfospreys.com/sports/womens-basketball/roster","Corrección de nombre: 'Georgi', no 'Georgina'"],
["Sol Depetris","F",TIER1,"Freshman","Base","N/D","Unión Florida","Bowling Green","NCAA D-I","Confirmada en roster 2026-27","N/D","https://bgsufalcons.com/sports/womens-basketball/roster/sol-depetris/20096",""],
["Paula López","F",TIER1,"N/D","Base / Escolta","N/D","San Blas Alicante (España)","Detroit Mercy","NCAA D-I","Confirmada por incorporación de mayo 2026","13/05/2026","https://detroittitans.com/news/2026/5/13/womens-basketball-detroit-mercy-womens-basketball-adds-transfer-paula-lopez.aspx","Vínculo selección U19 argentina; game notes la ubican España/Argentina"],
["Sol Castro","F",TIER1,"Graduate student","N/D","Río Colorado","N/D","San Francisco","NCAA D-I","Confirmada en roster 2026-27","N/D","https://usfdons.com/sports/womens-basketball/roster",""],
["Felipe Minzer","M",TIER2,"High school","Escolta / Alero","N/D","Zaragoza (España)","The Academy of Central Florida -> Austin Peay","NCAA D-I (incoming)","Confirmado como prospect con destino NCAA","10/04/2026","https://247sports.com/player/felipe-minzer-46151647/","Compromiso con Austin Peay reportado por Basquet Plus"],
["Máximo Adams","M",TIER2,"High school","Alero","N/D","Sierra Canyon HS (USA)","North Carolina","NCAA D-I (incoming)","Confirmado en high school estadounidense, firmado con UNC","19/11/2025","https://goheels.com/news/2025/11/19/mens-basketball-tar-heels-sign-maximo-adams","FIBA lo incluyó en roster tracker de Argentina en junio 2026"],
["Juan Sebastián Gorosito","M",TIER3,"Senior","Base / Escolta","Ceres, Santa Fe","N/D","San Diego","NCAA D-I","Senior 2025-26","N/D","https://usdtoreros.com/sports/mens-basketball/roster/juanse-gorosito/12922",""],
["Bautista Giralt","M",TIER3,"Redshirt Sophomore","Pívot","Buenos Aires","Obras Basket","Albany","NCAA D-I","Redshirt Sophomore 2025-26","N/D","https://ualbanysports.com/sports/mens-basketball/roster/bautista-giralt/11023",""],
["Lucas Mercandino","M",TIER3,"Junior","Alero","Córdoba","N/D","Lubbock Christian","NCAA D-II","Junior 2025-26","N/D","https://lcuchaps.com/sports/mens-basketball/roster/lucas-mercandino/4926",""],
["Ignacio García","M",TIER3,"Junior","Base","Villaguay, Entre Ríos","N/D","Lubbock Christian","NCAA D-II","Junior 2025-26","N/D","https://lcuchaps.com/sports/mens-basketball/roster/ignacio-garcia/4923",""],
["Facundo Aranda","M",TIER3,"Junior","Base","Mendoza","N/D","Biola","NCAA D-II","Junior 2025-26","N/D","https://athletics.biola.edu/sports/mens-basketball/roster/facundo-aranda/8456",""],
["Bautista Rodríguez","M",TIER3,"Sophomore","Alero","Santa Fe","N/D","Indiana Univ. of Pennsylvania","NCAA D-II","Sophomore 2025-26","N/D","https://iupathletics.com/sports/mens-basketball/roster/bautista-rodriguez/10579",""],
["Guido Wanschelbaum","M",TIER3,"Junior","Pívot","Buenos Aires","N/D","South Dakota Mines","NCAA D-II","Junior","N/D","https://gorockers.com/sports/mens-basketball/roster/guido-wanschelbaum/16885",""],
["Matías Zanotto","M",TIER3,"Freshman","Ala-pívot","Bahía Blanca","N/D","Augusta","NCAA D-II","Freshman 2025-26","N/D","https://augustajags.com/sports/mens-basketball/roster/matias-zanotto/5675",""],
["Tobías Lábaque","M",TIER3,"Sophomore","Escolta","Córdoba","N/D","Lynchburg","NCAA D-III","Sophomore 2025-26","N/D","https://lynchburgsports.com/sports/mens-basketball/roster/tobias-labaque/8439",""],
["Bernardo Zappia","M",TIER3,"Freshman","Pívot","Funes","N/D","Whittier","NCAA D-III","Freshman 2025-26","N/D","https://wcpoets.com/sports/mens-basketball/roster/bernardo-zappia/8378",""],
["Delfina Cergneux","F",TIER3,"Sophomore","N/D","Entre Ríos","N/D","Florida SouthWestern","NJCAA","Sophomore 2025-26","N/D","https://www.fswbucs.com/sports/wbkb/2025-26/roster","Nombre correcto: 'Cergneux' (no 'Cergneaux')"],
["Francisco Farabello","M",TIER4,"N/D (ex TCU)","Base","N/D","N/D","Creighton","NCAA D-I","No verificado en este informe estricto","N/D","https://www.basquetplus.com/ncaa-estados-unidos-arranca-temporada-argentinos-francisco-farabello-francisco-caffaro-rafael-martinez-lucas-mercandino-juanse-gorosito-felipe-palazzo","Pendiente confirmar roster 2026-27"],
["Felipe Palazzo","M",TIER4,"N/D","Pívot / Ala-pívot","N/D","Tucumán Básquet","Oregon State","NCAA D-I","No verificado en este informe estricto","N/D","https://www.basquetplus.com/ncaa-estados-unidos-arranca-temporada-argentinos-francisco-farabello-francisco-caffaro-rafael-martinez-lucas-mercandino-juanse-gorosito-felipe-palazzo","Pendiente confirmar roster 2026-27"],
["Stéfano Alesso","M",TIER4,"N/D","Pívot","Salta / Oberá","N/D","Stetson","NCAA D-I","No verificado en este informe estricto","N/D","https://basquetplus.com/ncaa-stefano-alesso-universidad-stetson-historia-pivote-gigante-salta-italia-obera-promesa-seleccion-argentina","Pendiente confirmar roster 2026-27"],
["Alonso Easterling","M",TIER4,"N/D","Ala-pívot / Pívot","N/D (vive en USA desde los 4 años)","N/D","Canisius (a confirmar)","NCAA D-I","No verificado en este informe estricto","N/D","https://basquetplus.com/ncaa-estados-unidos-alonso-easterling-nuevo-jugador-universidad-gardner-webb","Fuentes mencionan también Gardner-Webb"],
["Máximo Milovich","M",TIER4,"N/D (transfer)","N/D","N/D","N/D","UC San Diego","NCAA D-I","No verificado en este informe estricto","N/D","https://www.basquetplus.com/ncaa-maximo-milovich-division-I-uc-san-diego-transferencia",""],
["Francisco Cáffaro","M",TIER4,"N/D","Pívot","N/D","N/D","Virginia / Santa Clara","NCAA D-I","Carrera NCAA probablemente finalizada","N/D","https://en.wikipedia.org/wiki/Francisco_C%C3%A1ffaro","Verificar si sigue activo en NCAA o ya es profesional"],
["Juan Fernández","M",TIER4,"N/D","Base / Escolta","N/D","N/D","South Carolina (desde 2026-27)","NCAA D-I","Comprometido, aún no debuta","N/D","https://basquetplus.com/ncaa-nil-juan-fernandez-girona-south-carolina-dinero-lee-aaliya","Era profesional en Girona; caso atípico, no amateur tradicional"],
["Marcus Adams Jr.","M",TIER5,"N/D","Alero","N/D","N/D","Arizona State","NCAA D-I","Descartado","N/D","https://thesundevils.com/sports/mens-basketball/roster/player/marcus-adams","Fuente oficial de ASU lo presenta como jugador de Torrance, California, sin referencia a Argentina"],
["Brenda Fontana","F",TIER5,"Senior 2024-25","Ala-pívot","N/D","Vélez Sarsfield","Old Dominion","NCAA D-I","Sin afiliación nueva 2025-26/2026-27","N/D","https://odusports.com/sports/womens-basketball/roster/season/2024-25","Última evidencia fuerte: senior 2024-25"],
["Natalia Tondi","F",TIER5,"Senior 2024-25","Base","N/D","Unión Florida","Virginia Commonwealth","NCAA D-I","Sin continuidad en roster 2025-26","N/D","https://vcuathletics.com/sports/womens-basketball/roster/natalia-tondi/5593","Ya no figura en roster 2025-26 de VCU"],
["Tiziana Huici","F",TIER5,"Senior 2024-25","Base / Escolta","Tomás de Rocamora","N/D","Queens (NC)","NCAA D-II","Sin afiliación nueva posterior","N/D","https://queensathletics.com/sports/womens-basketball/roster/2024-2025","Tuvo senior night en febrero 2025"],
["Julia Paoletta","F",TIER5,"Senior 2024-25","N/D","Buenos Aires","N/D","Newberry","NCAA D-II","Sin continuidad pública posterior","N/D","https://newberrywolves.com/sports/womens-basketball/roster/2024-25",""],
]

tier_colors = {
    TIER1: "C6EFCE",
    TIER2: "FFEB9C",
    TIER3: "BDD7EE",
    TIER4: "F2F2F2",
    TIER5: "F4CCCC",
}

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF", name="Arial")
    c.fill = PatternFill("solid", start_color="1F4E78")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for r, row in enumerate(data, 2):
    tier = row[2]
    fill = PatternFill("solid", start_color=tier_colors.get(tier, "FFFFFF"))
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

widths = [22,8,32,16,16,16,28,30,14,28,14,40,36]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 32
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"

wb.save("Jugadores_Argentinos_USA_v2.xlsx")
print("saved", len(data), "filas")
